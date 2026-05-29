"""将安全尽调知识库Excel数据导入安全合规知识库，去重处理。"""

import openpyxl
import sys
import re
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

KB_ROOT = Path(r"D:\桌面20250227\工作\安全合规运营\2026年\安全合规运营管理\安全合规知识库\安全合规知识库")
EXCEL_PATH = Path(r"D:\桌面20250227\工作\安全合规运营\2026年\安全合规运营管理\安全合规知识库\安全尽调-知识库建设\安全尽调场景下知识库all.xlsx")

# Excel分类 → 知识库目录映射
CATEGORY_MAP = {
    "1. 组织与制度": {"dir": "01-制度体系", "subdir": None},
    "2. 数据安全": {"dir": "02-技术基线", "subdir": "03-数据安全基线"},
    "3. 网络安全": {"dir": "02-技术基线", "subdir": "05-网络安全基线"},
    "4. 主机与应用安全": {"dir": "02-技术基线", "subdir": "01-主机安全基线"},
    "5. 物理与环境安全": {"dir": "02-技术基线", "subdir": "06-物理环境安全基线"},
    "6. 业务连续性与灾备": {"dir": "06-应急响应", "subdir": "05-业务连续性"},
    "7. 应急响应": {"dir": "06-应急响应", "subdir": None},
    "8. 合规与审计": {"dir": "03-合规框架", "subdir": None},
    "9. 供应商与第三方管理": {"dir": "01-制度体系", "subdir": None},
    "10. 日志与审计": {"dir": "04-审计与整改", "subdir": None},
    "11. 移动安全与办公终端": {"dir": "02-技术基线", "subdir": "07-终端安全基线"},
    "12. 云安全管理": {"dir": "02-技术基线", "subdir": "02-云平台安全基线"},
}

# 子分类 → 文件名映射（用于合并到已有FAQ文件）
SUBCATEGORY_FILE_MAP = {
    # 01-制度体系
    "1.1 安全组织": ("07-FAQ", "01-安全制度FAQ.md", "安全组织"),
    "1.2 安全制度": ("07-FAQ", "01-安全制度FAQ.md", "安全制度"),
    "1.3 人员管理": ("07-FAQ", "01-安全制度FAQ.md", "人员管理"),
    "1.4 安全审计": ("07-FAQ", "01-安全制度FAQ.md", "安全审计"),
    # 02-数据安全
    "2.1 数据分类分级": ("07-FAQ", "02-数据安全FAQ.md", "数据分类分级"),
    "2.2 数据采集": ("07-FAQ", "02-数据安全FAQ.md", "数据采集"),
    "2.3 数据传输": ("07-FAQ", "02-数据安全FAQ.md", "数据传输"),
    "2.4 数据存储": ("07-FAQ", "02-数据安全FAQ.md", "数据存储"),
    "2.5 数据使用": ("07-FAQ", "02-数据安全FAQ.md", "数据使用"),
    "2.6 数据输出/共享": ("07-FAQ", "02-数据安全FAQ.md", "数据输出/共享"),
    "2.7 数据删除/销毁": ("07-FAQ", "02-数据安全FAQ.md", "数据删除/销毁"),
    "2.8 数据安全监控": ("07-FAQ", "02-数据安全FAQ.md", "数据安全监控"),
    "2.9 数据标准与质量": ("07-FAQ", "02-数据安全FAQ.md", "数据标准与质量"),
    "2.10 数据保护": ("07-FAQ", "02-数据安全FAQ.md", "数据保护"),
    # 03-网络安全
    "3.1 网络隔离": ("07-FAQ", "04-权限与访问FAQ.md", "网络访问控制"),
    "3.2 边界防护": ("07-FAQ", "04-权限与访问FAQ.md", "网络访问控制"),
    "3.3 访问控制": ("07-FAQ", "04-权限与访问FAQ.md", "数据访问权限"),
    "3.4 安全运维": ("07-FAQ", "04-权限与访问FAQ.md", "网络访问控制"),
    "3.5 网络准入": ("07-FAQ", "04-权限与访问FAQ.md", "网络访问控制"),
    # 04-主机与应用安全
    "4.1 主机安全": ("07-FAQ", "04-权限与访问FAQ.md", "身份鉴别"),
    "4.2 应用安全": ("07-FAQ", "02-数据安全FAQ.md", "数据使用"),
    "4.3 开发安全": ("07-FAQ", "02-数据安全FAQ.md", "数据使用"),
    "4.4 身份鉴别": ("07-FAQ", "04-权限与访问FAQ.md", "身份鉴别"),
    # 07-应急响应
    "7.1 应急预案": ("07-FAQ", "05-应急响应FAQ.md", "应急预案"),
    "7.2 应急演练": ("07-FAQ", "05-应急响应FAQ.md", "应急演练"),
    "7.3 事件通报": ("07-FAQ", "05-应急响应FAQ.md", "事件通报"),
    # 08-合规与审计
    "8.1 合规资质": ("07-FAQ", "03-合规要求FAQ.md", "合规资质"),
    "8.2 内部审计": ("07-FAQ", "03-合规要求FAQ.md", "内部审计"),
    "8.3 外部审计": ("07-FAQ", "03-合规要求FAQ.md", "外部审计"),
    "8.4 监管报送": ("07-FAQ", "03-合规要求FAQ.md", "监管报送"),
    # 09-供应商
    "9.1 供应商准入": ("07-FAQ", "01-安全制度FAQ.md", "安全制度"),
    "9.2 合同约束": ("07-FAQ", "01-安全制度FAQ.md", "安全制度"),
    "9.3 外包人员管理": ("07-FAQ", "01-安全制度FAQ.md", "人员管理"),
    # 10-日志与审计
    "10.1 日志留存": ("07-FAQ", "02-数据安全FAQ.md", "数据使用"),
    "10.2 日志审计": ("07-FAQ", "02-数据安全FAQ.md", "数据使用"),
    "10.3 时钟同步": ("07-FAQ", "02-数据安全FAQ.md", "数据使用"),
}


def read_existing_faq(filepath: Path) -> dict:
    """读取已有FAQ文件，提取所有问题用于去重。"""
    if not filepath.exists():
        return {}
    content = filepath.read_text(encoding='utf-8')
    questions = {}
    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("**Q:") or stripped.startswith("**Q："):
            q = stripped.lstrip("*").lstrip("Q").lstrip(":：").rstrip("**").strip()
            questions[q] = True
    return questions


def format_qa_item(seq: str, question: str, answer: str, standard: str = "", attachment: str = "") -> str:
    """格式化一条QA为markdown。"""
    lines = [f"**Q: {question}**", ""]
    a_parts = [f"A: {answer}"]
    if standard:
        a_parts.append(f"（标准/最佳实践：{standard}）")
    if attachment:
        a_parts.append(f"（证据：{attachment}）")
    lines.append(" ".join(a_parts))
    return "\n".join(lines)


def main():
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Read all rows
    rows_data = []
    last_cat = ""
    last_sub = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat = row[0] or last_cat
        sub = row[1] or last_sub
        last_cat = cat
        last_sub = sub
        seq = row[2] or ""
        question = row[3] or ""
        answer = row[4] or ""
        standard = row[5] or ""
        attachment = row[6] or ""
        permission = row[7] or ""
        if question:
            rows_data.append({
                "cat": cat, "sub": sub, "seq": str(seq),
                "question": question, "answer": answer,
                "standard": standard, "attachment": attachment,
                "permission": permission,
            })
    wb.close()

    # Group by target FAQ file
    file_groups = {}  # key: (faq_dir, faq_file), value: {section: [qa_items]}
    new_file_items = {}  # items that go to new files (not FAQ)

    for item in rows_data:
        sub = item["sub"]
        if sub in SUBCATEGORY_FILE_MAP:
            faq_dir, faq_file, section = SUBCATEGORY_FILE_MAP[sub]
            key = (faq_dir, faq_file)
            if key not in file_groups:
                file_groups[key] = {}
            if section not in file_groups[key]:
                file_groups[key][section] = []
            file_groups[key][section].append(item)
        else:
            # Goes to a new dedicated file
            cat = item["cat"]
            if cat not in new_file_items:
                new_file_items[cat] = []
            new_file_items[cat].append(item)

    # Process FAQ files: read existing, deduplicate, append new
    stats = {"added": 0, "skipped_dup": 0, "new_files": 0, "updated_files": 0}

    for (faq_dir, faq_file), sections in file_groups.items():
        filepath = KB_ROOT / faq_dir / faq_file
        existing_qs = read_existing_faq(filepath)
        existing_content = ""
        if filepath.exists():
            existing_content = filepath.read_text(encoding='utf-8')

        new_sections = []
        for section_name, items in sections.items():
            qa_lines = []
            for item in items:
                q = item["question"]
                # Deduplicate: check if question already exists (fuzzy match)
                q_clean = q.rstrip("？?").strip()
                is_dup = False
                for eq in existing_qs:
                    eq_clean = eq.rstrip("？?").strip()
                    if q_clean == eq_clean or q_clean in eq_clean or eq_clean in q_clean:
                        is_dup = True
                        break
                if is_dup:
                    stats["skipped_dup"] += 1
                    continue
                qa_text = format_qa_item(
                    item["seq"], item["question"], item["answer"],
                    item["standard"], item["attachment"],
                )
                qa_lines.append(qa_text)
                stats["added"] += 1

            if qa_lines:
                new_sections.append(f"\n## {section_name}\n\n" + "\n\n---\n\n".join(qa_lines))

        if new_sections:
            # Append to existing file
            with open(filepath, 'a', encoding='utf-8') as f:
                f.write("\n\n---\n".join(new_sections))
            stats["updated_files"] += 1

    # Process new files (categories without FAQ mapping)
    NEW_FILE_MAP = {
        "5. 物理与环境安全": ("02-技术基线", "06-物理环境安全基线", "物理与环境安全FAQ.md"),
        "6. 业务连续性与灾备": ("06-应急响应", "05-业务连续性", "业务连续性与灾备FAQ.md"),
        "11. 移动安全与办公终端": ("02-技术基线", "07-终端安全基线", "终端安全FAQ.md"),
        "12. 云安全管理": ("02-技术基线", "02-云平台安全基线", "云安全FAQ.md"),
    }

    for cat, items in new_file_items.items():
        if cat not in NEW_FILE_MAP:
            # Default: put in FAQ
            safe_name = re.sub(r'[^\w]', '_', cat)
            target = ("07-FAQ", None, f"{safe_name}FAQ.md")
        else:
            target = NEW_FILE_MAP[cat]

        top_dir, sub_dir, filename = target
        if sub_dir:
            filepath = KB_ROOT / top_dir / sub_dir / filename
        else:
            filepath = KB_ROOT / top_dir / filename

        filepath.parent.mkdir(parents=True, exist_ok=True)

        # Group by subcategory
        sub_groups = {}
        for item in items:
            sub = item["sub"]
            if sub not in sub_groups:
                sub_groups[sub] = []
            sub_groups[sub].append(item)

        lines = [f"# {cat}\n"]
        for sub_name, sub_items in sub_groups.items():
            lines.append(f"\n## {sub_name}\n")
            for item in sub_items:
                qa = format_qa_item(
                    item["seq"], item["question"], item["answer"],
                    item["standard"], item["attachment"],
                )
                lines.append(qa)
                lines.append("\n---\n")
                stats["added"] += 1

        filepath.write_text("\n".join(lines), encoding='utf-8')
        stats["new_files"] += 1

    print(f"导入完成:")
    print(f"  新增: {stats['added']} 条")
    print(f"  去重跳过: {stats['skipped_dup']} 条")
    print(f"  更新已有文件: {stats['updated_files']} 个")
    print(f"  新建文件: {stats['new_files']} 个")


if __name__ == "__main__":
    main()
